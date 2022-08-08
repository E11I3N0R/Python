import uuid
from datetime import datetime
from openpyxl import load_workbook, Workbook

# The Item class is the primary object created when we load in the inventory spreadsheet.
# It holds variables for all fields that a line of the inventory has.
# It contains functions for updating the quantities and removing quantities from the object.

class item:
    def __init__(self, var_upc, var_descr, var_maxqty, var_thresh, var_replen, var_onhand, var_unitpx, var_order):
        self.upc = var_upc  # UPC
        self.descr = var_descr  # Description
        self.maxqty = var_maxqty  # Item_Max_Qty
        self.thresh = var_thresh  # Order_Threshold
        self.replen = var_replen  # replenishment_order_qty
        self.onhand = var_onhand  # Item_on_hand
        self.unitpx = var_unitpx  # Unit price
        self.order = var_order  # Order_placed

    def print_check(self):
        y = [self.upc, self.descr, self.maxqty, self.thresh, self.replen, self.onhand, self.unitpx, self.order]
        print(y)

    def update_onhand_add(self, quantity):
        self.onhand += quantity

    def update_onhand_remove(self, quantity):
        if self.onhand < quantity:
            returnval = self.onhand
            print("\tInsufficient stock. Only added " + str(self.onhand) + " item/s.")
            self.onhand = 0
            return returnval
        else:
            self.onhand -= quantity
            return quantity

    def set_onhand(self, quantity):
        self.onhand = quantity

    def set_order_more(self):
        self.order = True

# This is the primary operating object for the store.
# It contains an overall inventory variable as well as running receipts for today's sales and a user.
# It has functions that handle all of the backend functions that directly reference inventory and all receipts.
class store:
    def __init__(self):
        self.inventory = {}
        self.today_receipts = []
        self.user = []
        self.inventory_titles = ["UPC", "Description", "Item_Max_Qty", "Order_Threshold",
                                 "replenishment_order_qty", "Item_on_hand", "Unit price", "Order_placed"]

    # Necessary for running any inventory actions.
    def load_data(self, ws):
        for row in ws.iter_rows(min_row=2, max_col=8, max_row=1039, values_only=True):
            temp = item(str(row[0]), row[1], row[2], row[3], row[4], row[5], row[6], row[7])
            self.inventory.update({temp.upc: temp})
        return self.inventory

    # Core function to look up inventory by UPC
    def get_item(self, get_upc):
        if get_upc in self.inventory.keys():
            return self.inventory.get(get_upc)
        else:
            return None

    def update_inventory_store(self, inv):
        self.inventory = inv

    def get_receipt(self, receipt_id):
        for x in self.today_receipts:
            if str(x.id) == receipt_id:
                return x

    def print_inventory(self):
        for x in self.inventory:
            print("UPC: " + str(self.inventory[x].upc) + " Description: " + self.inventory[
                x].descr + " Item_Max_Qty: " + str(self.inventory[x].maxqty))
            print("\tOrder_Threshold: " + str(self.inventory[x].thresh) + " replenishment_order_qty: " + str(
                self.inventory[x].replen))
            print("\tItem_on_hand: " + str(self.inventory[x].onhand) + " Unit price: " + str(
                self.inventory[x].unitpx) + " Order_placed: " + str(self.inventory[x].order) + "\n")

    # Calculates the sales for the day
    def today_sales(self):
        total = 0
        for x in self.today_receipts:
            print("Receipt ID: " + str(x.id))
            x.print_receipt()
            total += x.total_cash()
        print("Total Sales: " + str(round(total, 2)))

    # Creates a new inventory workbook using the hard coded column names and creates a new name using datetime
    # to prevent open user errors. Needs Zip Function to move through the inventory dictionary and column list together.
    def create_eod_inventory(self):
        workbook = Workbook()
        new_sheet = "ItemData_" + str(datetime.now()).replace(':', '').replace(' ', '_')[0:21]
        new_book = new_sheet + ".xlsx"
        ws = workbook.active
        ws.title = new_sheet
        i = 0
        for col in ws.iter_cols(min_row=1, max_col=8, max_row=1):
            for cell in col:
                cell.value = self.inventory_titles[i]
                i += 1
        inv_int = 2
        for item in self.inventory:
            for col, value in zip(ws.iter_cols(min_row=inv_int, max_col=8, max_row=inv_int), self.inventory[item].__dict__.items()):
                for cell in col:
                    cell.value = value[1]
            inv_int += 1
        workbook.save(new_book)
        return new_book

    # Check if inventories are below necessary thresholds and sets them to re-order.
    def reorder_inventory(self):
        new_file = "OrderItems_" + str(datetime.now()).replace(':', '').replace(' ', '_')[0:21] + ".txt"
        with open(new_file, "w") as file_write:
            for check_item in self.inventory:
                if self.inventory[check_item].onhand <= self.inventory[check_item].thresh:
                    if self.inventory[check_item].onhand + self.inventory[check_item].replen < self.inventory[check_item].maxqty:
                        self.inventory[check_item].set_order_more()
                        file_write.write(str(self.inventory[check_item].upc) + " " + str(self.inventory[check_item].descr) + " new quantity: " + str(self.inventory[check_item].onhand + self.inventory[check_item].replen) + "\n")
                    else:
                        self.inventory[check_item].set_order_more()
                        file_write.write(str(self.inventory[check_item].upc) + " " + str(self.inventory[check_item].descr) + " new quantity: " + str(self.inventory[check_item].maxqty) + "\n")
        return new_file

    # Creates or updates the running sales document
    def document_sales(self):
        new_file = "Running_Sales_Register.txt"
        with open(new_file, "a") as file_write:
            total = 0
            for x in self.today_receipts:
                file_write.write("Receipt ID: " + str(x.id) + " , ")
                x.print_receipt()
                total += x.total_cash()
            file_write.write("Total Sales: " + str(round(total, 2)) + " \n")
        return new_file

# Receipt class used to store all purchase transactions. Does not record returns just subtracts / removes items
class receipt:

    def __init__(self):
        self.id = uuid.uuid4()
        self.items_added = []

    def get_item(self, upc):
        for x in self.items_added:
            if x[0].upc == upc:
                return x

    def remove_item(self, upc):
        for x in self.items_added:
            if x[0].upc == upc:
                self.items_added.remove(x)

    def total_cash(self):
        total_cost = 0
        for item_qty in self.items_added:
            total_cost += (item_qty[0].unitpx * item_qty[1])
        return round(total_cost, 2)

    def print_receipt(self):
        print("Receipt id: " + str(self.id))
        for x in self.items_added:
            print(str(x[0].upc) + "  " + x[0].descr + "  " + str(x[0].unitpx) + "  " + str(x[1]) + "\n")

# Primary transaction class Inherits from the store for inventory functions like look ups. Used for sales mostly but can handle single returns. Tracks the working receipt and a copy
# of the dictionary for temporary purposes in case an order is canceled otherwise it's working dictionary will update
# the main store one.
class transaction(store):

    def __init__(self, inv):
        super().__init__()
        self.inventory = inv
        self.receipt = receipt()

    def sell_item(self, item_id, amount):
        item_sale = self.get_item(item_id)
        x = item_sale.update_onhand_remove(amount)
        self.receipt.items_added.append([item_sale, x])

    def return_item(self, item_id, amount):
        item_sale = self.get_item(item_id)
        receipt_item = self.receipt.get_item(item_sale.upc)
        if receipt_item:
            if receipt_item[1] > amount:
                receipt_item[1] -= amount
                item_sale.update_onhand_add(amount)
            else:
                self.receipt.remove_item(item_sale.upc)
                item_sale.update_onhand_add(amount)
                print("Removed " + item_sale.descr + " from receipt.")

    def complete_sale(self):
        return self.receipt

    def update_inventory(self):
        return self.inventory


# An expansion of regular transactions for returns.
class return_transaction(transaction):

    def __init__(self, receipt_obj, inv):
        super().__init__(inv)
        self.receipt = receipt_obj

    def return_all(self):  # return all items from this receipt id to inventory
        print("Returning the following: ")
        self.receipt.print_receipt()
        for x in self.receipt.items_added:
            x[0].update_onhand_add(x[1])
            print("Removed " + x[0].descr + " from receipt.")
        self.receipt.items_added.clear()


# MAIN run logic
user_store = store()
wb = load_workbook("RetailStoreItemData.xlsx", data_only=True)
sheet1 = wb["RetailStoreItemData"]
user_store.load_data(sheet1)

print("Welcome to the POS System\n")

attempt = 0
# User Login Loop
while attempt < 3:
    userid = input("Please enter userid: ")
    password = input("Please enter password: ")
    if userid != "EllieByler" or password != "mypassword":
        attempt += 1
        print("Invalid login. Please reenter credentials:")
    else:
        print("Login successful.")
        break
else:
    print("Too many incorrect login attempts. System locked.")
    exit()
# Main Run loop
run = True
while run:
    print("""\n*** MAIN MENU ***
    \n1 = New Sale, 2 = Return Item/s, 3 = Backroom Operations, 9 = Exit Application""")

    menu = True

    option = input("Please select your option: ")
    if option == "1":
        print("\n*** NEW SALE ***")
        new_transaction = transaction(user_store.inventory)
        while menu:
            print("\n1 = Add Item, 2 = Cancel Item/s, 7 = Complete Sale, 9 = Cancel")
            option = input("Please select your option: ")

            if option == "1":
                upc = input("Please enter the UPC to purchase: ")
                item_info = user_store.get_item(upc)
                if item_info:
                    print("\tItem: ", item_info.descr)
                    print("\tPrice: ", item_info.unitpx)
                    qty = int(input("\tPlease enter quantity: "))
                    new_transaction.sell_item(item_info.upc, qty)
                else:
                    print("\tUPC is not valid.")
            elif option == "2":
                upc = input("\tPlease enter the UPC to return: ")
                item_info = user_store.get_item(upc)
                if item_info:
                    if new_transaction.receipt.get_item(item_info.upc):
                        print("\tYou are returning " + item_info.descr)
                        qty = int(input("\tPlease enter quantity: "))
                        new_transaction.return_item(item_info.upc, qty)
                    else:
                        print("\tItem not on receipt.")
                else:
                    print("\tUPC is not valid.")
            elif option == "7":
                print("Sale completed.\n")
                x = new_transaction.complete_sale()
                # adds new receipt for today's totals
                user_store.today_receipts.append(x)
                # update working inventory from transaction
                user_store.update_inventory_store(new_transaction.update_inventory())
                x.print_receipt()
                print("Your total is", x.total_cash())
                menu = False
            elif option == "9":
                print("Cancel")
                menu = False
            else:
                print("Please select valid input.")
    elif option == "2":
        print("\n*** RETURN ITEM/S ***")
        receipt_id = input("Please enter receipt number: ")
        receipt_lookup = user_store.get_receipt(receipt_id)
        if receipt_lookup:
            new_return_transaction = return_transaction(receipt_lookup, user_store.inventory)
            while menu:
                print("\n1 = Return Single Item, 2 = Return All Items, 7 = Complete Returns, 9 = Cancel")
                option = input("Please select your option: ")
                if option == "1":
                    print("Return Single Item:")
                    upc = input("\tPlease enter the UPC to return: ")
                    item_info = user_store.get_item(upc)
                    if item_info:
                        if new_return_transaction.receipt.get_item(item_info.upc):
                            print("\tYou are returning " + item_info.descr)
                            qty = int(input("\tPlease enter quantity: "))
                            new_return_transaction.return_item(item_info.upc, qty)
                        else:
                            print("\tItem not on receipt.")
                    else:
                        print("\tUPC is not valid.")
                elif option == "2":
                    print("Return All Items:")
                    x = input("\tAre you sure you want to return all items? Y=yes, N=no: ")
                    if x == "Y" or x == "y":
                        new_return_transaction.return_all()
                        x = new_return_transaction.complete_sale()
                        if len(x.items_added) == 0: # removes receipt
                            user_store.today_receipts.remove(x)
                        user_store.inventory = new_return_transaction.update_inventory()
                        menu = False
                    elif x == "N" or x == "n":
                        print("Selected: No")
                    else:
                        print("Invalid input.")
                elif option == "7":
                    x = new_return_transaction.complete_sale()
                    if len(x.items_added) == 0:
                        user_store.today_receipts.remove(x)
                    user_store.inventory = new_return_transaction.update_inventory()
                    menu = False
                elif option == "9":
                    print("Cancel")
                    menu = False
                else:
                    print("Please select valid input.")
        else:
            print("Receipt not found.")
    elif option == "3":
        print("\n*** BACKROOM OPERATIONS ***")
        while menu:
            print("\n1 = Create Orders for Replenishment, 2 = Print Inventory Report,")
            print("\t3 = Create Today's Item Sold Report, 4 = Receipt Lookup")
            print("\t5 = Create Updated Inventory File, 6 = Update Running Sales")
            print("\t9 = Cancel")
            option = input("Please select your option: ")
            if option == "1":
                print("Create Orders for Replenishment:")
                print("Order report file: " + user_store.reorder_inventory())
            elif option == "2":
                print("Print Inventory Report:")
                user_store.print_inventory()
            elif option == "3":
                print("Create Today's Item Sold Report:")
                user_store.today_sales()
            elif option == "4":
                print("Receipt Lookup:")  # utility functionality
                receipt_id = input("Please enter receipt number: ")
                receipt_lookup = user_store.get_receipt(receipt_id)
                if receipt_lookup:
                    receipt_lookup.print_receipt()
                else:
                    print("Receipt not found.")
            elif option == "5":
                print("Created updated inventory file: " + user_store.create_eod_inventory())
            elif option == "6":
                print("Added today's items to Sales File: \n")
                print("Created / Updated Sales File: " + user_store.document_sales())
            elif option == "9":
                print("Cancel")
                menu = False
            else:
                print("Please select valid input.")
    elif option == "9":
        print("Created updated inventory file: " + user_store.create_eod_inventory())
        print("Logging Off")
        run = False
    else:
        print("Please select valid input.")


# DATA CHANGE LOG
# updated duplicate UPCs with "DUPE" at the end
# changed $0 unit price items to $7.99
# Replaced commas with - in item description
